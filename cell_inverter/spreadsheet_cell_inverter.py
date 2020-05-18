#spreadsheet_cell_inverter.py

# openpyxl modeule helps to deal with excel files
import openpyxl

#Method to invert all cells in a workbook
def invertCells(filename):

   #load_workbook( ) is used when you have to access an MS Excel file in openpyxl module for some operation
    wb = openpyxl.load_workbook(filename)

   #grabs the active worksheet
    sheet = wb.active

   #Creating new sheet with  name as inverted_cells
    newSheet = wb.create_sheet(index=0, title='inverted_cells')

    for rowObj in sheet.rows:
        for cellObj in rowObj:
            colIndex = cellObj.column
            rowIndex = cellObj.row

            newSheet.cell(row=colIndex, column=rowIndex).value = cellObj.value

    wb.save('result_'+filename)

#Main method
if __name__ == "__main__":

    #Provided input file name to invert its cells values
    invertCells('Marks.xlsx')