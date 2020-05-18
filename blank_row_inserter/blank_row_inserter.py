#blank_row_inserter.py
#The program should insert n number of  blank rows into the spreadsheet at the specified row number in the sheet
import openpyxl

"""Method takes index means row in file to start insert,num_blanks means number of blanks rows to insert 
and filename means name of file to insert blank rows in it"""
def blankRowInserter(index, num_blanks, filename):

    ##load_workbook( ) is used when you have to access an MS Excel file in openpyxl module for some operation
    wb = openpyxl.load_workbook(filename)

    ##grabs the active worksheet
    sheet = wb.active
    rows = tuple(sheet.rows)

    for rowObj in rows[::-1]:
        for cellObj in rowObj:
            c = cellObj.column
            r = cellObj.row

            if r >= index and r < index + num_blanks:
                sheet.cell(row=r + num_blanks, column=c).value = cellObj.value
                sheet.cell(row=r, column=c).value = ''
            elif r >= index + num_blanks:
                sheet.cell(row=r + num_blanks, column=c).value = cellObj.value

    wb.save('result_' + filename)


if __name__ == "__main__":
    #index means row in file to start insert
    index=int(input("Please enter index value-"))

    #num_of_blanks means number of blanks rows to insert
    num_of_blanks = int(input("Please enter number of blank rows to insert-"))

    #file_name means name of file to insert blank rows in it
    file_name = input("Please enter file name-")

    blankRowInserter(index, num_of_blanks, file_name)
