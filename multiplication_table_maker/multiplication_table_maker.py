#multiplication_table_maker.py
"""Program takes a number N from the user and creates an N×N multiplication table in an Excel spreadsheet"""

# openpyxl modeule helps to deal with excel fi
import openpyxl
from openpyxl.styles import Font

#Method to take number from user & to create N*N multiplication tablr
def multiplicationTable(n, filename='multiplicationTable.xlsx'):

    # create excel file
    wb = openpyxl.Workbook()

    # create worksheet & grabs the active worksheet
    sheet = wb.active

    #For bold labels
    boldFont = Font(bold=True)

    # write row headers
    for i in range(1, n + 1):
        sheet.cell(row=i + 1, column=1).value = i
        sheet.cell(row=i + 1, column=1).font = boldFont

    # write column headers
    for i in range(1, n + 1):
        sheet.cell(row=1, column=i + 1).value = i
        sheet.cell(row=1, column=i + 1).font = boldFont

    # write multiplication table
    for row in range(1, n + 1):
        for col in range(1, n + 1):
            sheet.cell(row=row + 1, column=col + 1).value = row * col

    # save table
    wb.save(filename)

#Main method
if __name__ == "__main__":
    #Takes number from user
    n=int(input("Enter the number from user to create an N×N multiplication table-"))
    multiplicationTable(int(n))