# text_to_excel.py
"""A program to read in the contents of several text files and insert those contents into a spreadsheet, with
one line of text per row"""
import os

# openpyxl modeule helps to deal with excel files
import openpyxl


# Method to convert text files to columns in excel worksheet
def textToSheet(directory, filename):
    # openpyxl.Workbook() creates the workbook
    wb = openpyxl.Workbook()

    # Creating new sheet with  name as result
    wb.create_sheet(index=0, title='result')

    # grabs the active worksheet
    sheet = wb.active

    colIndex = 1

    # write text files as columns in worksheet
    for file in os.listdir(directory):

        # Checks for file names with txt extension
        if file.endswith('.txt'):
            rowIndex = 1
            with open(file) as f:
                for line in f:
                    sheet.cell(row=rowIndex, column=colIndex).value = line
                    rowIndex += 1
            colIndex += 1

    # Saves data in excel file
    wb.save(filename)


# Main method
if __name__ == "__main__":
    # Providing path of current folder & output file name
    textToSheet('.', 'txt_to_excel_output.xlsx')
